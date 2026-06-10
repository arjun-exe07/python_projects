from datetime import datetime 
import pandas as pd 
import os 
from openpyxl import load_workbook 


def get_positive_int(prompt, min_val=None):          
  while True:         
    try:             
      val = int(input(prompt))             
      if val <= 0:                 
        print("Value must be positive.")                 
        continue             
      return val         
    except ValueError:             
      print("Invalid input. Please enter a valid integer.")  
      


def get_positive_float(prompt, min_val=None):          
  while True:         
    try:             
      val = float(input(prompt))             
      if val <= 0:                 
        print("Value must be positive.")                 
        continue                
      return val         
    except ValueError:             
      print("Invalid input. Please enter a valid number.")  
      
      

def getIncome():     
    
  my_income = {} 
  #Enter Income 
  source_count = get_positive_int("Enter number of income sources you want to insert :", min_val=1)  

  for i in range(1,source_count+1):         
    source = input(f" Enter the #{i} source name :").strip()      
    amount = get_positive_float(f"  Enter amount for {source} : ")         
    my_income[source] = my_income.get(source,0.0) +amount     
    print(f"The total income is: {sum(my_income.values()) :.2f}\n")     
  return my_income  
  


def getExpense(num):     
  my_bud = {}
  my_exp = {}

  print("\nEnter your Budget ::--")     
  #Enter Budget #     
  for i in range(num) :         
    cat = input ("Enter the category :")         
    amount = get_positive_float(f"Enter amount for {cat} :")         
    my_bud[cat] = my_bud.get(cat, 0.0) + amount  

  #Enter actual expenses for same category     
  print("\nEnter your actual expense for same categories ::--")     
  for cat in  my_bud:                
    amount = get_positive_float(f"Enter amount for {cat} :")         
    my_exp[cat] =  my_exp.get(cat, 0.0) + amount  

  return my_bud , my_exp

#Combine getExpense and getNewExpense

def getNewExpense(my_bud , my_exp):     
  new_cat_count = get_positive_int("Enter number of new categories you want to insert :", min_val=1)           
  print("\nEnter your Budget ::--")     
  for i in range(1, new_cat_count + 1):         
    cat = input ("Enter the category :").strip()         
    amount = get_positive_float(f"Enter amount for {cat} :")         
    my_bud[cat] = my_bud.get(cat, 0.0) + amount 
         
  #Enter actual expenses for same category     
  print("\nEnter your actual expense for same categories ::--")    
  for i in range(1, new_cat_count + 1):         
    cat = input ("Enter the category :").strip()         
    amount = get_positive_float(f"Enter amount for {cat} :")         
    my_exp[cat] = my_exp.get(cat, 0.0) + amount  
  return my_bud, my_exp
      


def displayBudgetAndExpenses(my_bud,my_exp):     
  all_categories = sorted(set(my_bud) | set(my_exp))     
  if not all_categories:         
    print("\nNo budget or expense categories available.")         
    return      
  for category in all_categories:         
    budget = my_bud.get(category)         
    expense = my_exp.get(category)         
    print(f"\n{category} :")          
    if budget is None:             
      print("No budget defined for this category.")             
      print(f"Expenses recorded: {expense:.2f}")             
      continue     

    if expense is None: 
      print("No expense recorded for this category yet.")             
      print(f"Budget set: {budget:.2f}")             
      continue          
    

    if budget == expense:             
      print("Well done, you have done it in your budget but,")             
      print("you have used all your budget")         
      
    elif budget < expense:             
      print("you have spent more than your budget.")             
      print("SUGGESTION :")             
      print("Recognize if you are spending due to boredom, stress, or sadness (retail therapy). ")             
      print("Use the 3 6 9 rule of money  to safeguard the budget. ")             
      print("If overspending is due to high-interest debt, focus on repayment and avoid adding new charges. ")         
    else:             
      print("congratulations ! ! !")             
      print("You still have budget remaining ")   
      


def report(my_bud , my_exp, income):     
  total_budget = sum(my_bud.values())     
  total_expenses = sum(my_exp.values())     
  total_income = sum(income.values()) if isinstance(income, dict) else float(income)     
  savings = total_income - total_expenses     
  remaining_bud = total_budget - total_expenses              
  print("\n\nReport ::--")     
  print(f"Total Budget is: {total_budget}")     
  print(f"Total expenses is: {total_expenses}")      
  print(f"Remaining Budget is: {remaining_bud}")     
  print(f"Saving is: {savings}")      
  
  if total_expenses > total_budget :         
    print("\n\t WARNING !!! \nyour expenses exceed your budget goals.")      
    
  if savings > 0 :         
    print("\nCongratulations !!! , Your Savings are greater than zero")         
    print(f"SUGGESTION :")         
    print("Equity Mutual Funds & Stocks: Ideal for long-term growth (5+ years). ")         
    print("Fixed Deposits (FDs) & Corporate Bonds: Offer guaranteed, fixed returns, suitable for short-to-medium-term goals.")         
    print("Gold ,Gold ETFS or Sovereign Gold Bonds: Acts as a hedge against economic uncertainty.")         
    print("National Pension System (NPS): A government-backed, low-cost retirement plan that invests in a mix of equity and debt.")  
    


def excel_report(my_bud , my_exp):     
  remarks = {}
  all_categories = sorted(set(my_bud) | set(my_exp)) 

  for category in all_categories:        
    budget = my_bud.get(category, 0.0)         
    expense = my_exp.get(category, 0.0)          
    if budget == expense:             
      remarks[category] = "Well done, you have used all budget"         
    elif budget < expense:             
      remarks[category] = "Warning!!, You have used more than your budget"         
    else:             
      remarks[category] = "Congrats!!, You still have budget remaining"      

  # Excel file name     
  file_name = "Data.xlsx"  

  # Current month-year sheet name   
  sheet_name = datetime.now().strftime("%B-%Y")

  info = {
            "Date": [datetime.now().strftime("%d-%m-%Y")] * len(all_categories), 
            "Category": all_categories,         
            "Budget": [my_bud.get(category, 0.0) for category in all_categories],        
            "Expenses": [my_exp.get(category, 0.0) for category in all_categories],         
            "Remark": [remarks[category] for category in all_categories]     
          }      
  # Convert to DataFrame     
  new_df = pd.DataFrame(info)      
  # Check if file exists     
  if os.path.exists(file_name):          
    # Load workbook        
    book = load_workbook(file_name)         
     
    # Check if current month sheet exists         
    if sheet_name in book.sheetnames:              
      # Read existing sheet data             
      existing_df = pd.read_excel(file_name, sheet_name=sheet_name)              
      # Append new data             
      updated_df = pd.concat([existing_df, new_df], ignore_index=True)                          
      book.save(file_name)             
      # Write updated data             
      with pd.ExcelWriter(file_name,engine="openpyxl",mode="a", if_sheet_exists='replace') as writer:                 
        updated_df.to_excel(writer,sheet_name=sheet_name,index=False)          
    else:             
      # Create new month sheet             
      with pd.ExcelWriter(file_name,engine="openpyxl",mode="a") as writer:                 
        new_df.to_excel(writer,sheet_name=sheet_name,index=False)      
  else:         
    # Create new Excel file and first sheet         
    with pd.ExcelWriter(file_name,engine="openpyxl") as writer:             
      new_df.to_excel(writer,sheet_name=sheet_name,index=False)      
  print("\n\nExcel Report ::--")     
  print(new_df)     
  print(f"Data added successfully in sheet: {sheet_name}")       
        
#Main Program 
def main() -> None :
  print("Personal Finance Management System")
  print("----------------------------------")

  income = getIncome() #Function to get income from different sources.
  categories_count = get_positive_int("Enter number of categories you want to insert :", min_val=1) 

  my_budget , my_expense = getExpense(categories_count)  #get Expense function to get budget and expenses for each category  

  while True :      
    print("\nEnter 1:- Add new expenses :")     
    print("Enter 2:- Display budget and expenses :")     
    print("Enter 3:- View the report")     
    print("Enter 4:- View the report in excel format :")     
    print("Enter 5:- Exit")     
    
    ch = get_positive_int("Enter your choice :")  

    if ch == 1 :         
      my_budget, my_expense = getNewExpense(my_budget , my_expense)  
    
    elif ch == 2 :         
      displayBudgetAndExpenses(my_budget ,my_expense)      
      
    elif ch == 3 :         
      report(my_budget , my_expense, income)      
      
    elif ch == 4:         
      excel_report(my_budget , my_expense)      
    
    elif ch == 5 :         
      print("Thanks for using our finance monitoring system.")         
      break      
    else :         
      print("Invalid choice. \nPlease choice again")         
      